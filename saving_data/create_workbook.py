from openpyxl import Workbook


class DataWorkbook:
    def __init__(self):
        self.__workbook = Workbook()

    def save_data(self, data_base):
        sheet = self.__workbook.active
        sheet.title = 'Smartphones_Parameters'
        column_names = ['Full Name', 'Brand', 'Model', 'Rating', 'Standards', 'Weight', 'Screen', 'Battery', 'Charging',
                        'Memory', 'RAM', 'System', 'Processor', 'Market introduction']
        column = 1
        for column_name in column_names:
            sheet.cell(column=column, row=1, value=column_name)
            column = column + 1
        row = 2
        for model in data_base:
            sheet.cell(column=1, row=row, value=model)
            sheet.cell(column=2, row=row, value=data_base[model]['brand'])
            sheet.cell(column=3, row=row, value=data_base[model]['model'])
            sheet.cell(column=4, row=row, value=data_base[model]['rating'])
            sheet.cell(column=5, row=row, value=data_base[model]['standards'])
            sheet.cell(column=6, row=row, value=data_base[model]['weight'])
            sheet.cell(column=7, row=row, value=data_base[model]['screen'])
            sheet.cell(column=8, row=row, value=data_base[model]['battery'])
            sheet.cell(column=9, row=row, value=data_base[model]['charging'])
            sheet.cell(column=10, row=row, value=data_base[model]['memory'])
            sheet.cell(column=11, row=row, value=data_base[model]['memory_RAM'])
            sheet.cell(column=12, row=row, value=data_base[model]['system'])
            sheet.cell(column=13, row=row, value=data_base[model]['processor'])
            sheet.cell(column=14, row=row, value=data_base[model]['on_market'])

            row = row + 1
        self.__adjust_column_width(sheet)

    def save_collaborative_ranking(self, collaborative_ranking):
        sheet = self.__workbook.create_sheet(title='Collaborative_Filtering')
        sheet.cell(column=1, row=1, value='Place')
        sheet.cell(column=2, row=1, value='Smartphones')
        for place in collaborative_ranking:
            sheet.cell(column=1, row=int(place)+1, value=place)
            for model in collaborative_ranking[place]:
                index = collaborative_ranking[place].index(model)
                sheet.cell(column=2+index, row=int(place)+1, value=model)
        self.__adjust_column_width(sheet)

    def save_content_ranking(self):
        sheet = self.__workbook.create_sheet(title='Content_Based_Filtering')

    def save_file_workbook(self):
        self.__workbook.save('data.xlsx')

    @staticmethod
    def __adjust_column_width(sheet):
        dims = {}
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet.column_dimensions[col].width = value + 3
